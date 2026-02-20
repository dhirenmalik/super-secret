import pandas as pd
import numpy as np
import requests
import pickle
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
try:
    from thefuzz import fuzz, process
except ImportError:
    try:
        from fuzzywuzzy import fuzz, process
    except ImportError:
        # Fallback to rapidfuzz if available
        try:
            from rapidfuzz import fuzz, process
        except ImportError:
            # Re-raise the first one if everything fails
            from thefuzz import fuzz, process
import unicodedata
import json
from tqdm import tqdm


class BrandMatcher:
    def __init__(self, config):
        self.config = config
        self.model = SentenceTransformer(config.get('model_name', 'all-MiniLM-L6-v2'))
        self.model_name = config.get('model_name', 'all-MiniLM-L6-v2')
        self.threshold = config.get('threshold', 0.85)
        self.use_parent_lookup = config.get('use_parent_lookup', False)
        self.embedding_match_threshold = config.get('embedding_match_threshold', 0.9)
        self.fuzzy_match_threshold = config.get('fuzzy_match_threshold', 90)
        self.parent_match_threshold = config.get('parent_match_threshold', 90)
        # Removed private brands path and list
        self.cache_path = config.get('cache_path', 'parent_cache.pkl')

        self.Combine_flag_groups = {}
        self.parent_cache = {}
        self.historical_map = {}
        self.historical_brands = []
        self.historical_embeddings = None
        self.next_Combine_flag = 1
        self.loaded = False

        # Removed loading private brands
        self.load_cache_from_file()

    def load_historical_data(self, data, brand_col='UNIQUE_BRAND_NAME', flag_col='Combine_flag'):
        print(f"Loading historical data from {data}...")
        if isinstance(data, str):
            with open(data, 'r') as f:
                json_data = json.load(f)
            rows = []
            for idx, brand_list in enumerate(json_data.values(), start=1):
                for brand in brand_list:
                    rows.append({brand_col: brand, flag_col: idx})
            df = pd.DataFrame(rows)
        elif isinstance(data, pd.DataFrame):
            df = data.copy()
        else:
            raise TypeError("`data` must be a filepath string or a pandas DataFrame.")

        if flag_col not in df.columns:
            print(f"Flag column '{flag_col}' not found, initializing as NaN")
            df[flag_col] = np.nan

        self.historical_data = df
        self.brand_col = brand_col
        self.flag_col = flag_col

        print("Encoding embeddings for historical brands...")
        embeddings = self.model.encode(self.historical_data[brand_col].tolist(), show_progress_bar=True)
        self.historical_data['embedding'] = list(embeddings)
        self.historical_brands = self.historical_data[brand_col].tolist()
        self.historical_embeddings = embeddings

        self.historical_map = dict(zip(self.historical_data[brand_col], self.historical_data[flag_col]))

        self.loaded = True
        if not df[flag_col].dropna().empty:
            try:
                # Robustly find max integer flag
                numeric_flags = pd.to_numeric(df[flag_col], errors='coerce').dropna()
                if not numeric_flags.empty:
                    self.next_Combine_flag = int(numeric_flags.max()) + 1
                else:
                    self.next_Combine_flag = 1
            except (ValueError, TypeError):
                self.next_Combine_flag = 1
        else:
            self.next_Combine_flag = 1
        print(f"Historical data loaded. Next Combine flag: {self.next_Combine_flag}")

    def normalize_text(self, text):
        normalized = unicodedata.normalize('NFKD', text)
        return normalized.encode('ASCII', 'ignore').decode('utf-8').upper()

    def get_parent_company(self, brand_name):
        if brand_name in self.parent_cache:
            return self.parent_cache[brand_name]

        endpoint_url = "https://query.wikidata.org/sparql"
        query = f"""
            SELECT ?parentLabel WHERE {{
            ?brand_entity rdfs:label ?label.
            FILTER(LANG(?label) = "en" && LCASE(STR(?label)) = LCASE("{brand_name}")).
            ?brand_entity wdt:P749 ?parent.
            SERVICE wikibase:label {{ bd:serviceParam wikibase:language "[AUTO_LANGUAGE],en". }}
            }}
            LIMIT 1
        """
        headers = {
            "User-Agent": "BrandMatcher/1.0 (your@email.com)",
            "Accept": "application/json"
        }
        try:
            response = requests.get(endpoint_url, params={'query': query, 'format': 'json'}, headers=headers, timeout=10)
            response.encoding = 'utf-8'
            if response.status_code == 200:
                results = response.json().get("results", {}).get("bindings", [])
                if results:
                    parent = self.normalize_text(results[0]["parentLabel"]["value"])
                    self.parent_cache[brand_name] = parent
                    self.save_cache_to_file()
                    return parent
        except Exception as e:
            print(f"Error querying Wikidata for {brand_name}: {e}")

        self.parent_cache[brand_name] = None
        self.save_cache_to_file()
        return None

    def match_brands(self, new_df, brand_col='UNIQUE_BRAND_NAME'):
        print("Starting brand matching...")
        new_embeddings = self.model.encode(new_df[brand_col].tolist(), show_progress_bar=False)
        sim_matrix = cosine_similarity(new_embeddings, self.historical_embeddings)

        if 'Combine_flag' not in new_df.columns:
            new_df['Combine_flag'] = None
        if 'comment' not in new_df.columns:
            new_df['comment'] = ''

        for i, (idx, row) in enumerate(new_df.iterrows()):
            brand = row[brand_col]
            # Removed private brand check here

            best_match_idx = sim_matrix[i].argmax()
            best_score = sim_matrix[i][best_match_idx]
            matched_brand = self.historical_brands[best_match_idx]
            matched_flag = self.historical_map.get(matched_brand, None)

            if best_score > self.embedding_match_threshold and pd.notna(matched_flag):
                new_df.at[idx, 'Combine_flag'] = matched_flag
                new_df.at[idx, 'comment'] = f'historic brand/group matched with "{matched_brand}" (score: {best_score:.2f})'
                continue

            valid_brands = [b for b, flag in self.historical_map.items() if pd.notna(flag)]
            fuzzy_brand_match = process.extractOne(brand, valid_brands, scorer=fuzz.token_sort_ratio)
            if fuzzy_brand_match and fuzzy_brand_match[1] >= self.fuzzy_match_threshold:
                matched_brand, score = fuzzy_brand_match
                matched_flag = self.historical_map[matched_brand]
                new_df.at[idx, 'Combine_flag'] = matched_flag
                new_df.at[idx, 'comment'] = f'fuzzy matched brand "{brand}" ~ "{matched_brand}" (score: {score})'
                continue

            new_df.at[idx, 'Combine_flag'] = None
            new_df.at[idx, 'comment'] = ''

        flag_counts = new_df['Combine_flag'].value_counts(dropna=True)
        multi_flags = flag_counts[flag_counts > 1].index.tolist()

        for idx, row in new_df.iterrows():
            flag = row['Combine_flag']
            if pd.notna(flag) and flag not in multi_flags:
                new_df.at[idx, 'Combine_flag'] = None
                new_df.at[idx, 'comment'] = ''

        print("Brand matching completed.")
        return new_df

    def analyze_brands_from_excel(self, excel_path=None, input_sheet=None, output_sheet=None):
        excel_path = excel_path or self.config.get('input_excel_path')
        input_sheet = input_sheet or self.config.get('input_sheet', 'Sheet1')
        output_sheet = output_sheet or self.config.get('output_sheet', 'brand_analysis')

        if excel_path is None:
            raise ValueError("Excel input path must be provided as argument or in config['input_excel_path']")

        print(f"Loading Excel file: {excel_path} (sheet: {input_sheet})")
        try:
            df = pd.read_excel(excel_path, sheet_name=input_sheet)
            print(f"Loaded {len(df)} rows from {input_sheet}")
        except Exception as e:
            print(f"Failed to load input Excel file: {e}")
            return None

        print("Calculating Total Spend for each brand...")
        df["Total_Spend"] = (
            df.get("M_SEARCH_SPEND", pd.Series([0]*len(df))).fillna(0) +
            df.get("M_OFF_DIS_TOTAL_SUM_SPEND", pd.Series([0]*len(df))).fillna(0) +
            df.get("M_ON_DIS_TOTAL_SUM_SPEND", pd.Series([0]*len(df))).fillna(0)
        )

        print("Grouping data by UNIQUE_BRAND_NAME...")
        grouped_df = df.groupby("UNIQUE_BRAND_NAME", dropna=False).agg({
            "O_SALE": "sum",
            "O_UNIT": "sum",
            "Total_Spend": "sum"
        }).reset_index()

        grouped_df["Exclude_Flag"] = 0
        grouped_df["Combine_flag"] = None
        print(f"Start matching brands in grouped data ({len(grouped_df)})...")
        df_matched = self.match_brands(grouped_df, brand_col="UNIQUE_BRAND_NAME")

        print("Setting exclusion flags...")
        condition = (df_matched["O_SALE"].fillna(0) == 0) | (df_matched["Total_Spend"].fillna(0) == 0)
        for i in tqdm(range(len(df_matched)), desc="Exclusion Flag Processing", ncols=100):
            if condition[i] and pd.isna(df_matched.at[i, 'Combine_flag']):
                df_matched.at[i, "Exclude_Flag"] = 1
            elif not condition[i] and pd.notna(df_matched.at[i, 'Combine_flag']):
                df_matched.at[i, "Exclude_Flag"] = 0

        print("Normalizing Combine_flag values...")
        unique_flags = df_matched['Combine_flag'].dropna().unique()
        combine_map = {old: new for new, old in enumerate(unique_flags, start=1)}
        df_matched['Combine_flag'] = df_matched['Combine_flag'].map(combine_map)

        print("Calculating summary stats...")
        total_sales = df_matched["O_SALE"].sum()
        total_spends = df_matched["Total_Spend"].sum()
        total_units = df_matched["O_UNIT"].sum()
        included = df_matched[df_matched["Exclude_Flag"] == 0]
        excluded = df_matched[df_matched["Exclude_Flag"] == 1]

        summary_data = [
            ["Total", total_sales, total_spends, total_units],
            ["Included", included["O_SALE"].sum(), included["Total_Spend"].sum(), included["O_UNIT"].sum()],
            ["Excluded", excluded["O_SALE"].sum(), excluded["Total_Spend"].sum(), excluded["O_UNIT"].sum()]
        ]
        summary_df = pd.DataFrame(summary_data, columns=["After Exclude Flag Analysis", "Sales", "Spends", "Units"])

        print(f"Writing matched data and summary to sheet '{output_sheet}' in Excel: {excel_path}")

        from openpyxl import load_workbook
        book = load_workbook(excel_path)
        if output_sheet in book.sheetnames:
            std = book[output_sheet]
            book.remove(std)
        if f"{output_sheet}_summary" in book.sheetnames:
            std_sum = book[f"{output_sheet}_summary"]
            book.remove(std_sum)
        book.save(excel_path)

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a') as writer:
            df_matched.to_excel(writer, sheet_name=output_sheet, index=False)
            summary_df.to_excel(writer, sheet_name=f"{output_sheet}_summary", index=False)

        print("Data saved successfully.")


    def save_cache_to_file(self):
        print(f"Saving parent cache to {self.cache_path}...")
        try:
            with open(self.cache_path, 'wb') as cache_file:
                pickle.dump(self.parent_cache, cache_file)
            print("Cache saved successfully.")
        except Exception as e:
            print(f"Error saving cache: {e}")

    def load_cache_from_file(self):
        print(f"Loading cache from {self.cache_path}...")
        try:
            with open(self.cache_path, 'rb') as cache_file:
                self.parent_cache = pickle.load(cache_file)
            print("Cache loaded successfully.")
        except FileNotFoundError:
            print(f"No cache file found at {self.cache_path}, starting with empty cache.")
            self.parent_cache = {}
        except Exception as e:
            print(f"Error loading cache: {e}")
            self.parent_cache = {}
